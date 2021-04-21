from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle, Font, Alignment
from data.config import TABLES_FOLDER
from pathlib import Path
import re
import os

_file_prefix = 'result_'
_ext = 'xlsx'
_max_file_size = 10485760
_sheet_name = 'result'
_header_font = Font(name='Arial',
                    size=13,
                    bold=True)
_centred = Alignment(vertical='center',
                     horizontal='center')


def get_name(path=TABLES_FOLDER):
    if not path:
        path = Path(__file__).parent.parent.parent / 'tables'

    if not isinstance(Path, str):
        path = str(path)

    if not os.path.exists(path):
        os.mkdir(path)

    not_filtered = list(filter(lambda filename: filename if _file_prefix in filename else None, os.listdir(path)))
    filtered_files_names = list(filter(
        lambda filename: filename if os.path.getsize(os.path.join(path, filename)) < _max_file_size else None,
        not_filtered))

    if not filtered_files_names:
        try:
            return _file_prefix + \
                   str(max([int(re.findall('[0-9]+', filename)[-1]) for filename in os.listdir(path)]) + 1) + \
                   '.' + _ext
        except ValueError:
            return os.path.join(path, _file_prefix + '0' + '.' + _ext)
    else:
        return os.path.join(path, filtered_files_names[-1])


def write_to_exel(data: dict):
    filename = get_name()
    keys = list(data.keys())

    try:
        wb = load_workbook(get_name())
    except FileNotFoundError:
        wb = Workbook()

    try:
        ws: Worksheet = wb['result']
    except KeyError:
        ws: Worksheet = wb.create_sheet('result')

    try:
        last_row = ws._cells.get(list(ws._cells.keys())[-1]).row
    except IndexError:
        for i in range(1, len(data.keys()) + 1):
            cell = ws.cell(row=1, column=i, value=keys[i - 1])
            cell.font = _header_font
            cell.alignment = _centred
        last_row = 1

    for i in range(1, len(data) + 1):
        cell = ws.cell(row=last_row + 1, column=i, value=data.get(keys[i - 1]))
        cell.alignment = _centred
    wb.save(filename)

